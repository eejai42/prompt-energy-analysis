#!/usr/bin/env python3
"""
CMCC Truth Model Workbook Builder

Creates an Excel workbook that encodes a small CMCC-style model (S/D/L/A/F)
for the prompt about constructed truths vs reality-constrained invariants.

Run:
  python cmcc_truth_model_builder.py

Output:
  cmcc_truth_model.xlsx (created in the same directory as this script)

Dependencies:
  pip install openpyxl
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment


# -----------------------------
# Helpers (formatting + tables)
# -----------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")      # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL  = PatternFill("solid", fgColor="FFF2CC")      # light yellow
INPUT_FONT  = Font(color="0000FF")                        # blue inputs

THIN = Side(style="thin", color="9E9E9E")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP  = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header_row(ws, row: int = 1, max_col: int | None = None) -> None:
    if max_col is None:
        max_col = ws.max_column
    ws.row_dimensions[row].height = 22
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_data_range(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 18
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = WRAP


def add_table(ws, name: str, ref: str) -> None:
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


# -----------------------------
# Workbook builder
# -----------------------------

def build_workbook(out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # README
    ws = wb.create_sheet("README")
    ws["A1"] = "CMCC Truth Model (Constructed vs. Reality-Constrained Truth)"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "What this workbook is:"
    ws["A3"].font = Font(bold=True, size=12)
    ws["A4"] = (
        "A small CMCC-style model built with the five primitives:\n"
        "S = Schema (what columns exist)\n"
        "D = Data (rows)\n"
        "L = Lookups (references between tables)\n"
        "A = Aggregations (roll-ups / checks)\n"
        "F = Formulas (pure calculated fields)\n\n"
        "It encodes your prompt as computable rules about: (a) constructed truths "
        "(definitions/conventions) and (b) reality-constrained invariants "
        "(unit-independent quantities)."
    )

    ws["A6"] = "How to play with it:"
    ws["A6"].font = Font(bold=True, size=12)
    ws["A7"] = (
        "1) Go to D_Constants and change a value in a yellow cell (e.g., electron mass).\n"
        "2) Derived calculations in F_Calculations and checks in D_Claims update.\n"
        "3) D_Questions gives TRUE/FALSE answers to the prompt’s key questions.\n\n"
        "Yellow cells = inputs. Everything else is computed."
    )

    ws["A9"] = "Primary sources used for constants/conversions (also recorded as URLs in Notes/comments):"
    ws["A9"].font = Font(bold=True, size=12)
    ws["A10"] = "NIST CODATA constants: https://physics.nist.gov/constants"
    ws["A11"] = "BIPM SI Brochure: https://www.bipm.org/en/publications/si-brochure"
    ws["A12"] = "NIST unit converter (eV↔J): https://physics.nist.gov/cgi-bin/cuu/Convert?From=ev&To=j"

    ws["A14"] = "Tabs:"
    ws["A14"].font = Font(bold=True, size=12)
    ws["A15"] = "S_Schema (meta), D_Units, D_Constants, F_Calculations, D_Instances, D_Claims, D_Questions, A_Dashboard"

    ws.column_dimensions["A"].width = 120
    for r in range(4, 16):
        ws.row_dimensions[r].height = 40
    ws.freeze_panes = "A4"

    # S_Schema
    ws = wb.create_sheet("S_Schema")
    headers = ["Table", "Field", "Type", "Required?", "Description"]
    ws.append(headers)
    schema_rows = [
        ("D_Units", "UnitID", "Text", "Y", "Stable identifier (e.g., U_J, U_eV)"),
        ("D_Units", "ToSI_Mult", "Number", "Y", "Multiply by this to convert to SI base unit"),
        ("D_Units", "ToSI_Offset", "Number", "Y", "Add this to convert to SI base unit (temperature)"),
        ("D_Constants", "ConstantID", "Text", "Y", "Stable identifier (e.g., C_me, C_c)"),
        ("D_Constants", "Value", "Number", "Y", "Value in the given unit"),
        ("D_Constants", "Value_SI", "Formula", "Y", "Value converted to SI using D_Units"),
        ("F_Calculations", "CalcID", "Text", "Y", "Stable identifier for derived calculations"),
        ("F_Calculations", "Result_SI", "Formula", "Y", "Derived value in SI"),
        ("D_Instances", "InstanceID", "Text", "Y", "Scenario instance for a quantity"),
        ("D_Instances", "Canonical_SI", "Formula", "Y", "Instance value converted to SI"),
        ("D_Claims", "Pass", "Formula", "Y", "TRUE/FALSE whether the claim's criteria are met"),
        ("D_Questions", "Answer", "Formula", "Y", "TRUE/FALSE answer derived from claims"),
    ]
    for row in schema_rows:
        ws.append(list(row))
    set_col_widths(ws, [18, 18, 12, 12, 70])
    style_header_row(ws, 1, len(headers))
    style_data_range(ws, 2, ws.max_row, 1, len(headers))
    ws.freeze_panes = "A2"

    # D_Units
    ws_units = wb.create_sheet("D_Units")
    u_headers = ["UnitID", "UnitName", "QuantityKind", "SI_UnitName", "ToSI_Mult", "ToSI_Offset", "Notes_URL"]
    ws_units.append(u_headers)
    units_data = [
        ("U_J", "joule", "Energy", "J", 1.0, 0.0, "https://www.bipm.org/en/measurement-units"),
        ("U_eV", "electronvolt", "Energy", "J", 1.602176634e-19, 0.0, "https://physics.nist.gov/cgi-bin/cuu/Value?evj="),
        ("U_ftlbf", "foot-pound force", "Energy", "J", 1.3558179483314, 0.0, "https://www.convertunits.com/from/ft-lbf/to/joule"),
        ("U_K", "kelvin", "Temperature", "K", 1.0, 0.0, "https://www.bipm.org/documents/20126/41483022/SI-Brochure-9-concise-EN.pdf"),
        ("U_C", "degree Celsius", "Temperature", "K", 1.0, 273.15, "https://www.bipm.org/documents/20126/41483022/SI-Brochure-9-concise-EN.pdf"),
        ("U_kg", "kilogram", "Mass", "kg", 1.0, 0.0, "https://www.bipm.org/en/measurement-units"),
        ("U_mps", "metre per second", "Speed", "m/s", 1.0, 0.0, "https://www.bipm.org/documents/20126/41483022/SI-Brochure-9-EN.pdf"),
        ("U_Coul", "coulomb", "Charge", "C", 1.0, 0.0, "https://www.bipm.org/en/measurement-units"),
    ]
    for row in units_data:
        ws_units.append(list(row))
    set_col_widths(ws_units, [10, 20, 14, 12, 14, 14, 55])
    style_header_row(ws_units, 1, len(u_headers))
    style_data_range(ws_units, 2, ws_units.max_row, 1, len(u_headers))
    ws_units.freeze_panes = "A2"
    add_table(ws_units, "tblUnits", f"A1:{get_column_letter(len(u_headers))}{ws_units.max_row}")

    # We'll use fixed VLOOKUP ranges to keep compatibility with older Excel
    units_range = f"D_Units!$A$2:$G${1 + len(units_data)}"

    # D_Constants
    ws_c = wb.create_sheet("D_Constants")
    c_headers = [
        "ConstantID", "Name", "Symbol", "Value", "UnitID", "SourceLayer", "Authority_URL",
        "Unit_ToSI_Mult", "Unit_ToSI_Offset", "Value_SI", "Notes",
    ]
    ws_c.append(c_headers)
    constants_rows = [
        ("C_me", "electron mass", "m_e", 9.1093837139e-31, "U_kg", "Measured (CODATA)", "https://physics.nist.gov/cgi-bin/cuu/Value?me=", "", "", "", "From NIST CODATA"),
        ("C_c", "speed of light in vacuum", "c", 299792458.0, "U_mps", "Defined (SI)", "https://www.bipm.org/en/measurement-units", "", "", "", "Exact by SI definition"),
        ("C_e", "elementary charge", "e", 1.602176634e-19, "U_Coul", "Defined (SI)", "https://www.bipm.org/en/measurement-units", "", "", "", "Exact by SI definition"),
        ("C_mec2", "electron mass energy equivalent", "m_e c^2", 8.1871057880e-14, "U_J", "Measured/Adjusted (CODATA)", "https://physics.nist.gov/cgi-bin/cuu/Value?mec2%7Csearch_for=electron+mass", "", "", "", "Reference value to validate E=mc^2 calc"),
        ("C_abs0", "absolute zero (thermodynamic)", "0 K", 0.0, "U_K", "Defined (scale anchor)", "https://www.bipm.org/documents/20126/41483022/SI-Brochure-9-concise-EN.pdf", "", "", "", "0 K is absolute zero on Kelvin scale"),
    ]
    for row in constants_rows:
        ws_c.append(list(row))

    # Fill VLOOKUP formulas and Value_SI
    for r in range(2, ws_c.max_row + 1):
        ws_c.cell(r, 8).value = f'=VLOOKUP(E{r},{units_range},5,FALSE)'
        ws_c.cell(r, 9).value = f'=VLOOKUP(E{r},{units_range},6,FALSE)'
        ws_c.cell(r,10).value = f'=D{r}*H{r}+I{r}'

    set_col_widths(ws_c, [12, 28, 16, 16, 10, 22, 45, 14, 16, 16, 32])
    style_header_row(ws_c, 1, len(c_headers))
    style_data_range(ws_c, 2, ws_c.max_row, 1, len(c_headers))
    ws_c.freeze_panes = "A2"
    add_table(ws_c, "tblConstants", f"A1:{get_column_letter(len(c_headers))}{ws_c.max_row}")

    # Mark input cells (Value column) and add source comments
    for r in range(2, ws_c.max_row + 1):
        cell = ws_c.cell(r, 4)
        cell.fill = INPUT_FILL
        cell.font = INPUT_FONT
        cell.number_format = "0.0000000000E+00"
        src = ws_c.cell(r, 7).value
        cell.comment = Comment(f"Input cell. Source: {src}", "CMCC")
        ws_c.cell(r, 10).number_format = "0.0000000000E+00"

    # VLOOKUP range for constants (Value_SI is column 10)
    const_range = f"D_Constants!$A$2:$J${1 + len(constants_rows)}"

    def vconst(const_id: str) -> str:
        return f'VLOOKUP("{const_id}",{const_range},10,FALSE)'

    # F_Calculations
    ws_f = wb.create_sheet("F_Calculations")
    f_headers = ["CalcID", "Name", "Expression", "Result_UnitID", "Result_SI", "Notes"]
    ws_f.append(f_headers)
    calc_rows = [
        ("F_Ecalc", "Electron rest energy from E=mc^2", "m_e * c^2", "U_J", "", "Uses constants C_me and C_c"),
        ("F_eV_to_J", "eV to J conversion", "e (since 1 V = 1 J/C, so eV = e·J)", "U_J", "", "Checks constructed truth: conversion derives from defined constants"),
        ("F_T_C_from_K", "Celsius from Kelvin", "t(°C)=T(K)-273.15", "U_C", "", "Uses BIPM relationship"),
    ]
    for row in calc_rows:
        ws_f.append(list(row))

    # VLOOKUP range for calcs (Result_SI is column 5)
    calcs_range = f"F_Calculations!$A$2:$E${1 + len(calc_rows)}"

    def vcalc(calc_id: str) -> str:
        return f'VLOOKUP("{calc_id}",{calcs_range},5,FALSE)'

    ws_f["E2"] = f"={vconst('C_me')}*({vconst('C_c')}^2)"
    ws_f["E3"] = f"={vconst('C_e')}"
    # E4 intentionally left blank (demonstration formula is text in Expression)
    set_col_widths(ws_f, [12, 28, 40, 14, 18, 40])
    style_header_row(ws_f, 1, len(f_headers))
    style_data_range(ws_f, 2, ws_f.max_row, 1, len(f_headers))
    ws_f.freeze_panes = "A2"
    add_table(ws_f, "tblCalcs", f"A1:{get_column_letter(len(f_headers))}{ws_f.max_row}")
    ws_f["E2"].number_format = "0.0000000000E+00"
    ws_f["E3"].number_format = "0.0000000000E+00"

    # D_Instances
    ws_i = wb.create_sheet("D_Instances")
    i_headers = [
        "InstanceID", "Scenario", "QuantityName", "QuantityKind", "Value", "UnitID",
        "Unit_ToSI_Mult", "Unit_ToSI_Offset", "Canonical_SI", "Notes",
    ]
    ws_i.append(i_headers)

    instances_rows = [
        ("I_Ecalc_J", "ElectronRestEnergy", "E (calc)", "Energy", "", "U_J", "", "", "", "From F_Ecalc"),
        ("I_Ecalc_eV", "ElectronRestEnergy", "E (calc)", "Energy", "", "U_eV", "", "", "", "Converted from SI"),
        ("I_Ecalc_ftlbf", "ElectronRestEnergy", "E (calc)", "Energy", "", "U_ftlbf", "", "", "", "Converted from SI"),
        ("I_Eexpected_J", "ElectronRestEnergy", "E (CODATA)", "Energy", "", "U_J", "", "", "", "Reference constant C_mec2"),
        ("I_Tabs0_K", "AbsoluteZero", "T absolute", "Temperature", 0.0, "U_K", "", "", "", "0 Kelvin"),
        ("I_Tabs0_C", "AbsoluteZero", "T absolute", "Temperature", "", "U_C", "", "", "", "Celsius representation of absolute zero"),
    ]
    for row in instances_rows:
        ws_i.append(list(row))

    # Look up unit conversion factors
    for r in range(2, ws_i.max_row + 1):
        ws_i.cell(r, 7).value = f'=VLOOKUP(F{r},{units_range},5,FALSE)'
        ws_i.cell(r, 8).value = f'=VLOOKUP(F{r},{units_range},6,FALSE)'

    # Fill instance Value formulas
    ws_i["E2"] = f"={vcalc('F_Ecalc')}"
    ws_i["E3"] = f"={vcalc('F_Ecalc')}/{vcalc('F_eV_to_J')}"
    ws_i["E4"] = f"={vcalc('F_Ecalc')}/VLOOKUP(\"U_ftlbf\",{units_range},5,FALSE)"
    ws_i["E5"] = f"={vconst('C_mec2')}"
    ws_i["E7"] = f"=E6-273.15"

    # Canonical SI
    for r in range(2, ws_i.max_row + 1):
        ws_i.cell(r, 9).value = f"=E{r}*G{r}+H{r}"
        ws_i.cell(r, 5).number_format = "0.0000000000E+00"
        ws_i.cell(r, 9).number_format = "0.0000000000E+00"

    set_col_widths(ws_i, [14, 18, 18, 14, 18, 10, 14, 16, 18, 40])
    style_header_row(ws_i, 1, len(i_headers))
    style_data_range(ws_i, 2, ws_i.max_row, 1, len(i_headers))
    ws_i.freeze_panes = "A2"
    add_table(ws_i, "tblInstances", f"A1:{get_column_letter(len(i_headers))}{ws_i.max_row}")

    # D_Claims
    ws_cl = wb.create_sheet("D_Claims")
    cl_headers = [
        "ClaimID", "ClaimText", "ClaimType", "Criteria",
        "InstanceID_1", "InstanceID_2", "InstanceID_3",
        "ToleranceAbs_SI",
        "SI_1", "SI_2", "SI_3",
        "RangeAbs", "Pass", "TruthSource_Layer",
    ]
    ws_cl.append(cl_headers)

    claims_rows = [
        ("CL1", "Electron rest energy is invariant across units (J, eV, ft-lbf).", "Reality-constrained invariant",
         "Max(SI)-Min(SI) <= tolerance", "I_Ecalc_J", "I_Ecalc_eV", "I_Ecalc_ftlbf", 1e-25, "", "", "", "", "", ""),
        ("CL2", "E = m c^2 (using m_e and c) matches CODATA m_e c^2 within tolerance.", "Reality-constrained invariant",
         "Abs(SI_calc - SI_ref) <= tolerance", "I_Ecalc_J", "I_Eexpected_J", "", 1e-22, "", "", "", "", "", ""),
        ("CL3", "0 K and -273.15 °C represent the same absolute temperature.", "Mixed (scale + reality)",
         "Abs(SI_K - SI_C) <= tolerance", "I_Tabs0_K", "I_Tabs0_C", "", 1e-9, "", "", "", "", "", ""),
        ("CL4", "1 eV equals e joules (constructed via SI definition of e and volt).", "Constructed (definition/convention)",
         "Abs(eV_to_J - e) <= tolerance", "", "", "", 0.0, "", "", "", "", "", ""),
    ]
    for row in claims_rows:
        ws_cl.append(list(row))

    inst_range = f"D_Instances!$A$2:$I${1 + len(instances_rows)}"  # Canonical_SI is col 9

    # Fill formulas for SI lookups, range, pass, and truth-source heuristic
    for r in range(2, ws_cl.max_row + 1):
        ws_cl.cell(r, 9).value  = f'=IF(E{r}="","",VLOOKUP(E{r},{inst_range},9,FALSE))'
        ws_cl.cell(r,10).value  = f'=IF(F{r}="","",VLOOKUP(F{r},{inst_range},9,FALSE))'
        ws_cl.cell(r,11).value  = f'=IF(G{r}="","",VLOOKUP(G{r},{inst_range},9,FALSE))'
        ws_cl.cell(r,12).value  = f'=IF(K{r}="",ABS(I{r}-J{r}),MAX(I{r},J{r},K{r})-MIN(I{r},J{r},K{r}))'
        ws_cl.cell(r,13).value  = f"=L{r}<=H{r}"
        ws_cl.cell(r,14).value  = f'=IF(LEFT(C{r},11)="Constructed","Convention/Definition","Reality (constraint) + Convention (coordinates)")'
        for col in (8, 9, 10, 11, 12):
            ws_cl.cell(r, col).number_format = "0.0000000000E+00"

    # Patch CL4: compare calculation F_eV_to_J vs constant C_e
    r_cl4 = 5  # header row 1, CL4 is row 5
    ws_cl.cell(r_cl4, 9).value  = f"={vcalc('F_eV_to_J')}"
    ws_cl.cell(r_cl4,10).value  = f"={vconst('C_e')}"
    ws_cl.cell(r_cl4,11).value  = ""
    ws_cl.cell(r_cl4,12).value  = f"=ABS(I{r_cl4}-J{r_cl4})"
    ws_cl.cell(r_cl4,13).value  = f"=L{r_cl4}<=H{r_cl4}"

    set_col_widths(ws_cl, [10, 55, 28, 32, 14, 14, 14, 16, 16, 16, 16, 16, 10, 28])
    style_header_row(ws_cl, 1, len(cl_headers))
    style_data_range(ws_cl, 2, ws_cl.max_row, 1, len(cl_headers))
    ws_cl.freeze_panes = "A2"
    add_table(ws_cl, "tblClaims", f"A1:{get_column_letter(len(cl_headers))}{ws_cl.max_row}")

    # D_Questions
    ws_q = wb.create_sheet("D_Questions")
    q_headers = ["QID", "Question", "Mapped_Claims", "Answer", "Explanation"]
    ws_q.append(q_headers)

    questions_rows = [
        ("Q0", "Are some truths in the model constructed (true because defined)?", "CL4", "",
         "TRUE if the definitional/conventional check (CL4) passes."),
        ("Q1", "Are there consistent truths that don’t come from construction (i.e., unit-independent invariants)?", "CL1,CL2,CL3", "",
         "TRUE if at least one reality-constrained invariant claim passes."),
        ("Q2", "Does changing the measurement system (J vs ft-lbf vs eV) change the underlying electron rest energy?", "CL1", "",
         "FALSE if CL1 passes (energy is invariant after conversion)."),
        ("Q3", "Does E=mc^2 hold as long as appropriate conversions are done?", "CL2", "",
         "TRUE if CL2 passes."),
        ("Q4", "Is 0 K the same temperature as −273.15 °C (after conversion)?", "CL3", "",
         "TRUE if CL3 passes."),
        ("Q5", "Is 'reality' the constraining source of truth for invariants in this model?", "CL1,CL2,CL3", "",
         "TRUE if at least one invariant passes; conventions set coordinates (units), reality constrains what fits."),
    ]
    for row in questions_rows:
        ws_q.append(list(row))

    claims_range = f"D_Claims!$A$2:$M${1 + len(claims_rows)}"  # Pass is column 13
    def vclaim_pass(cid: str) -> str:
        return f'VLOOKUP("{cid}",{claims_range},13,FALSE)'

    ws_q["D2"] = f"={vclaim_pass('CL4')}"
    ws_q["D3"] = f"=OR({vclaim_pass('CL1')},{vclaim_pass('CL2')},{vclaim_pass('CL3')})"
    ws_q["D4"] = f"=NOT({vclaim_pass('CL1')})"
    ws_q["D5"] = f"={vclaim_pass('CL2')}"
    ws_q["D6"] = f"={vclaim_pass('CL3')}"
    ws_q["D7"] = f"=OR({vclaim_pass('CL1')},{vclaim_pass('CL2')},{vclaim_pass('CL3')})"

    set_col_widths(ws_q, [8, 72, 18, 10, 55])
    style_header_row(ws_q, 1, len(q_headers))
    style_data_range(ws_q, 2, ws_q.max_row, 1, len(q_headers))
    ws_q.freeze_panes = "A2"
    add_table(ws_q, "tblQuestions", f"A1:{get_column_letter(len(q_headers))}{ws_q.max_row}")

    # A_Dashboard
    ws_d = wb.create_sheet("A_Dashboard")
    ws_d["A1"] = "Dashboard"
    ws_d["A1"].font = Font(bold=True, size=16)
    ws_d["A3"] = "Key computed values (SI):"
    ws_d["A3"].font = Font(bold=True)

    ws_d["A4"] = "Electron rest energy (calc) [J]"
    ws_d["B4"] = f"={vcalc('F_Ecalc')}"
    ws_d["A5"] = "Electron rest energy (CODATA) [J]"
    ws_d["B5"] = f"={vconst('C_mec2')}"
    ws_d["A6"] = "Absolute zero [K]"
    ws_d["B6"] = f"={vconst('C_abs0')}"

    ws_d["A8"] = "Claims passing"
    ws_d["A8"].font = Font(bold=True)
    ws_d["B8"] = f'=COUNTIF(D_Claims!$M$2:$M${1 + len(claims_rows)},TRUE)'
    ws_d["A9"] = "Claims total"
    ws_d["B9"] = f"={len(claims_rows)}"

    ws_d["A11"] = "Questions (TRUE/FALSE):"
    ws_d["A11"].font = Font(bold=True)
    for idx, r in enumerate(range(2, 2 + len(questions_rows)), start=0):
        dr = 12 + idx
        ws_d[f"A{dr}"] = f"=D_Questions!B{r}"
        ws_d[f"B{dr}"] = f"=D_Questions!D{r}"
        ws_d[f"B{dr}"].alignment = CENTER

    set_col_widths(ws_d, [70, 18, 18, 18])
    for r in range(4, 7):
        ws_d[f"B{r}"].number_format = "0.0000000000E+00"

    for r in range(4, 10):
        for c in range(1, 3):
            ws_d.cell(r, c).border = BORDER
            ws_d.cell(r, c).alignment = WRAP
    for r in range(12, 12 + len(questions_rows)):
        for c in range(1, 3):
            ws_d.cell(r, c).border = BORDER
            ws_d.cell(r, c).alignment = WRAP

    ws_d.freeze_panes = "A4"

    # Final polish
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    wb.save(out_path)


def main() -> None:
    out_path = Path(__file__).with_name("cmcc_truth_model.xlsx")
    build_workbook(out_path)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
